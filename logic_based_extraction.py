#!/usr/bin/env python3
"""
Logic-based Text Extraction for All OU Companies Klarna Invoices
Processes all klarna txt files and generates comprehensive Excel output
Supports 8 OU companies: AUSTRALIA, UK, IRELAND, TOWERS, STYLES_SERVICES, CORPORATION, US_SERVICES, CANADA
"""

import os
import re
import pandas as pd
from pathlib import Path


def extract_shein_australia_data(lines):
    """Extract data for SHEIN DISTRIBUTION AUSTRALIA PTY LIMITED"""
    return extract_data_by_company(lines, "AUSTRALIA")


def extract_shein_uk_data(lines):
    """Extract data for SHEIN DISTRIBUTION UK LIMITED"""
    return extract_data_by_company(lines, "UK")


def extract_infinite_styles_ireland_data(lines):
    """Extract data for INFINITE STYLES ECOMMERCE CO., LIMITED"""
    return extract_data_by_company(lines, "IRELAND")


def extract_infinite_towers_data(lines):
    """Extract data for INFINITE TOWERS SERVICES LIMITED"""
    return extract_data_by_company(lines, "TOWERS")


def extract_infinite_styles_services_data(lines):
    """Extract data for INFINITE STYLES SERVICES CO., LIMITED"""
    return extract_data_by_company(lines, "STYLES_SERVICES")


def extract_shein_corporation_data(lines):
    """Extract data for SHEIN DISTRIBUTION CORPORATION"""
    return extract_data_by_company(lines, "CORPORATION")


def extract_shein_us_services_data(lines):
    """Extract data for SHEIN US Services, LLC"""
    return extract_data_by_company(lines, "US_SERVICES")


def extract_shein_canada_data(lines):
    """Extract data for Shein Distribution Canada Limited"""
    return extract_data_by_company(lines, "CANADA")


def detect_ou_company(lines):
    """根据第8行内容检测OU公司"""
    if len(lines) >= 8:
        line_8 = lines[7].strip()
        separator = "     "  # 5 spaces
        if separator in line_8:
            return line_8.split(separator)[0].strip()
    return "未知公司"


def detect_company_type(ou_company):
    """根据OU公司名称检测公司类型"""
    if "AUSTRALIA" in ou_company:
        return "AUSTRALIA"
    elif "UK" in ou_company:
        return "UK"
    elif "INFINITE STYLES ECOMMERCE" in ou_company:
        return "IRELAND"
    elif "INFINITE TOWERS" in ou_company:
        return "TOWERS"
    elif "INFINITE STYLES SERVICES" in ou_company:
        return "STYLES_SERVICES"
    elif "SHEIN DISTRIBUTION CORPORATION" in ou_company:
        return "CORPORATION"
    elif "SHEIN US Services" in ou_company:
        return "US_SERVICES"
    elif "Shein Distribution Canada" in ou_company:
        return "CANADA"
    else:
        return "UNKNOWN"


def extract_data_by_company(lines, company_type):
    """
    Extract data based on company type with all fixes applied
    """
    result = {
        'invoice_number': '',
        'our_company_name': '',
        'our_company_address': '',
        'our_tax_id': '',
        'invoice_date': '',
        'net_amount': '',
        'tax_rate': '',
        'tax_amount': '',
        'total_amount': '',
        'currency': '',
        'vendor_name': '',
        'vendor_address': '',
        'vendor_tax_id': '',
        'filename': '',
        'processing_errors': []
    }

    try:
        # 发票号码：第5行右侧连续字符串
        if len(lines) > 4:
            line_5 = lines[4].strip()
            if '     ' in line_5:
                result['invoice_number'] = line_5.split('     ')[-1].strip()
            else:
                result['invoice_number'] = line_5.strip()

        # 检测OU公司并设置对应的公司名
        ou_company = detect_ou_company(lines)
        result['our_company_name'] = ou_company

        # 我方公司地址：根据公司类型处理
        if len(lines) > 8:
            line_9 = lines[8].strip()

            if company_type == "IRELAND":
                # ISEL: 直到", IE VAT ID"前的字符串
                if ', IE VAT ID' in line_9:
                    result['our_company_address'] = line_9.split(', IE VAT ID')[0].strip()
                else:
                    result['our_company_address'] = line_9

            elif company_type == "TOWERS":
                # Towers: 直到连续5个空格的字符串
                if '     ' in line_9:
                    result['our_company_address'] = line_9.split('     ')[0].strip()
                else:
                    result['our_company_address'] = line_9

            elif company_type == "STYLES_SERVICES":
                # Styles Services: 直到", IE VAT ID"前的字符串
                if ', IE VAT ID' in line_9:
                    result['our_company_address'] = line_9.split(', IE VAT ID')[0].strip()
                else:
                    result['our_company_address'] = line_9

            elif company_type == "CORPORATION":
                # Corporation: 直到", US VAT ID"前的字符串
                if ', US VAT ID' in line_9:
                    result['our_company_address'] = line_9.split(', US VAT ID')[0].strip()
                else:
                    result['our_company_address'] = line_9

            elif company_type == "US_SERVICES":
                # US Services: 直到", US  VAT ID"前的字符串
                if ', US  VAT ID' in line_9:
                    result['our_company_address'] = line_9.split(', US  VAT ID')[0].strip()
                else:
                    result['our_company_address'] = line_9

            elif company_type == "CANADA":
                # Canada: 直到"GST/HST/QST number:"前的字符串
                if 'GST/HST/QST number:' in line_9:
                    result['our_company_address'] = line_9.split('GST/HST/QST number:')[0].strip()
                else:
                    result['our_company_address'] = line_9

            else:  # AUSTRALIA, UK
                # AUSTRALIA和UK：直到连续5个空格的字符串
                if '     ' in line_9:
                    result['our_company_address'] = line_9.split('     ')[0].strip()
                else:
                    result['our_company_address'] = line_9

        # 我方税号：根据公司类型使用不同的标识符
        if company_type == "AUSTRALIA":
            # 查找包含ABN的行
            for line in lines:
                if 'ABN' in line:
                    abn_match = re.search(r'ABN[:\s]+([^\s]+)', line)
                    if abn_match:
                        result['our_tax_id'] = abn_match.group(1)
                    break

        elif company_type in ["UK", "TOWERS"]:
            # 查找包含VAT ID的行
            for line in lines:
                if 'VAT ID:' in line:
                    vat_match = re.search(r'VAT ID:\s*([^\s]+)', line)
                    if vat_match:
                        result['our_tax_id'] = vat_match.group(1)
                    break

        elif company_type in ["IRELAND", "STYLES_SERVICES"]:
            # 查找第9行中的IE VAT ID
            if len(lines) > 8:
                line_9 = lines[8]
                ie_vat_match = re.search(r'IE VAT ID:\s*([^\s]+)', line_9)
                if ie_vat_match:
                    result['our_tax_id'] = ie_vat_match.group(1)

        elif company_type in ["CORPORATION", "US_SERVICES"]:
            # 查找第9行中的US VAT ID
            if len(lines) > 8:
                line_9 = lines[8]
                us_vat_match = re.search(r'US\s*VAT ID:\s*([^\s]+)', line_9)
                if us_vat_match:
                    result['our_tax_id'] = us_vat_match.group(1)

        elif company_type == "CANADA":
            # 查找包含GST/HST/QST number的行 - 第9行右侧
            if len(lines) > 8:
                line_9 = lines[8]
                gst_match = re.search(r'GST/HST/QST number:\s*([^\s]+)', line_9)
                if gst_match:
                    result['our_tax_id'] = gst_match.group(1)

        # 发票日期：第10行 Payout date:后的 dd MMM YYY 字符串
        if len(lines) > 9:
            line_10 = lines[9]
            date_match = re.search(r'Payout date:\s*(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})', line_10)
            if date_match:
                result['invoice_date'] = date_match.group(1)

        # 不含税金额：根据公司类型查找Fees行
        fees_keyword = 'Fees 1' if company_type in ["AUSTRALIA", "UK", "TOWERS", "IRELAND", "STYLES_SERVICES"] else 'Fees '

        for line in lines:
            if fees_keyword in line:
                # 查找 "Transactions" 后的数字
                transactions_match = re.search(r'Transactions[^0-9]*([\d,]+\.\d{2})', line)
                if transactions_match:
                    amount_str = transactions_match.group(1).replace(',', '')
                    result['net_amount'] = float(amount_str)
                break

        # 税率和税额：根据公司类型处理
        if company_type == "AUSTRALIA":
            tax_keyword = 'GST on fees'
            result['tax_rate'] = '10.00%'
            tax_amount = 0

        elif company_type in ["UK", "TOWERS"]:
            tax_keyword = 'VAT on fees'
            result['tax_rate'] = '20.00%'
            tax_amount = 0

        elif company_type in ["IRELAND", "STYLES_SERVICES", "CORPORATION", "US_SERVICES", "CANADA"]:
            result['tax_rate'] = '0%'
            result['tax_amount'] = 0.0
            tax_keyword = None

        else:
            tax_keyword = None
            tax_amount = 0

        # 处理有税的公司
        if tax_keyword:
            found_tax = False
            for line in lines:
                if tax_keyword in line:
                    found_tax = True
                    rate_match = re.search(r'\(([^)]*%[^)]*)\)', line)
                    if rate_match:
                        result['tax_rate'] = rate_match.group(1).strip()

                    negative_amount_match = re.search(r'\)[^-]*-([\d,]+\.\d{2})', line)
                    if negative_amount_match:
                        amount_str = negative_amount_match.group(1).replace(',', '')
                        tax_amount = float(amount_str)
                    else:
                        after_paren = re.search(r'\)[^0-9-]*([\d,]+\.\d{2})', line)
                        if after_paren:
                            amount_str = after_paren.group(1).replace(',', '')
                            tax_amount = float(amount_str)
                    break

            if not found_tax:
                tax_amount = 0.0

            result['tax_amount'] = tax_amount

        # 含税金额：Total costs and fees行中右侧负号开始的数字
        for line in lines:
            if 'Total costs and fees' in line:
                total_match = re.search(r'-[\d,]+\.\d{2}', line)
                if total_match:
                    total_str = total_match.group(0).replace('-', '').replace(',', '')
                    result['total_amount'] = float(total_str)
                else:
                    positive_match = re.search(r'[\d,]+\.\d{2}', line)
                    if positive_match:
                        total_str = positive_match.group(0).replace(',', '')
                        result['total_amount'] = float(total_str)
                break

        # 币种：Payout 且不是 Payout date 行中的3位ISO币种
        for line in lines:
            if 'Payout' in line and 'Payout date' not in line:
                currency_match = re.search(r'\b[A-Z]{3}\b', line)
                if currency_match:
                    result['currency'] = currency_match.group(0)
                break

        # 供应商信息：Need support行下一行 - 应用所有修复
        need_support_found = False
        for i, line in enumerate(lines):
            if 'Need support' in line:
                need_support_found = True
                if i + 1 < len(lines):
                    next_line = lines[i + 1]

                    # 供应商名称：第1个","前的字符串
                    if ',' in next_line:
                        result['vendor_name'] = next_line.split(',')[0].strip()

                    # 供应商地址和税号：根据公司类型处理 - 应用所有修复
                    if company_type == "AUSTRALIA":
                        if ',' in next_line and '• ABN' in next_line:
                            first_comma_pos = next_line.find(',')
                            abn_marker_pos = next_line.find('• ABN')
                            if first_comma_pos != -1 and abn_marker_pos != -1:
                                address = next_line[first_comma_pos + 1:abn_marker_pos].strip()
                                result['vendor_address'] = address

                        if 'ABN' in next_line:
                            abn_part = next_line.split('ABN')[1].strip()
                            result['vendor_tax_id'] = abn_part

                    elif company_type in ["UK", "TOWERS"]:
                        # 供应商地址：第1个","和"• VAT numbers"中间的所有字符
                        if ',' in next_line and '• VAT numbers' in next_line:
                            first_comma_pos = next_line.find(',')
                            vat_marker_pos = next_line.find('• VAT numbers')
                            if first_comma_pos != -1 and vat_marker_pos != -1:
                                address = next_line[first_comma_pos + 1:vat_marker_pos].strip()
                                result['vendor_address'] = address

                        # 🔧 修复：供应商税号 - 优先提取GB开头的税号，避免提取SE税号
                        vendor_tax_id = ''

                        # 首先在同一行查找VAT numbers后的GB税号
                        if 'VAT numbers' in next_line:
                            vat_numbers_part = next_line.split('VAT numbers')[1]
                            # 优先查找GB开头的税号
                            gb_match = re.search(r'(GB[^,\s]*)', vat_numbers_part)
                            if gb_match:
                                vendor_tax_id = gb_match.group(1).strip()

                        # 如果同一行没有找到，检查下一行
                        if not vendor_tax_id and i + 2 < len(lines):
                            line_after_next = lines[i + 2]
                            if 'GB' in line_after_next:
                                gb_match = re.search(r'(GB[^,\s]*)', line_after_next)
                                if gb_match:
                                    vendor_tax_id = gb_match.group(1).strip()

                        result['vendor_tax_id'] = vendor_tax_id

                    elif company_type in ["IRELAND", "STYLES_SERVICES"]:
                        if ',' in next_line and '• VAT numbers' in next_line:
                            first_comma_pos = next_line.find(',')
                            vat_marker_pos = next_line.find('• VAT numbers')
                            if first_comma_pos != -1 and vat_marker_pos != -1:
                                address = next_line[first_comma_pos + 1:vat_marker_pos].strip()
                                result['vendor_address'] = address

                        vendor_tax_id = ''
                        if 'Registration number' in next_line:
                            reg_part = next_line.split('Registration number')[1].strip()
                            if i + 2 < len(lines):
                                next_line_after = lines[i + 2].strip()
                                five_space_pos = next_line_after.find('     ')
                                if five_space_pos != -1:
                                    next_line_after = next_line_after[:five_space_pos].strip()
                                if reg_part and next_line_after:
                                    vendor_tax_id = f"{reg_part}{next_line_after}"
                                elif reg_part:
                                    vendor_tax_id = reg_part
                                else:
                                    vendor_tax_id = next_line_after
                        result['vendor_tax_id'] = vendor_tax_id

                    elif company_type in ["CORPORATION", "US_SERVICES"]:
                        if ',' in next_line and '• TIN' in next_line:
                            first_comma_pos = next_line.find(',')
                            tin_marker_pos = next_line.find('• TIN')
                            if first_comma_pos != -1 and tin_marker_pos != -1:
                                address = next_line[first_comma_pos + 1:tin_marker_pos].strip()
                                result['vendor_address'] = address

                        if '• TIN' in next_line:
                            tin_part = next_line.split('• TIN')[1].strip()
                            result['vendor_tax_id'] = tin_part

                    elif company_type == "CANADA":
                        # 供应商地址：第1个","和"• GST/HST/QST"中间的所有字符
                        if ',' in next_line and '• GST/HST/QST' in next_line:
                            first_comma_pos = next_line.find(',')
                            gst_marker_pos = next_line.find('• GST/HST/QST')
                            if first_comma_pos != -1 and gst_marker_pos != -1:
                                address = next_line[first_comma_pos + 1:gst_marker_pos].strip()
                                result['vendor_address'] = address

                        # 🔧 修复：供应商税号 - 加拿大特定逻辑，提取完整税号格式
                        vendor_tax_id = ''

                        # 检查下第2行中 "number " 后的内容
                        if i + 2 < len(lines):
                            third_line = lines[i + 1]
                            fourth_line = lines[i + 2]

                            # 优先从第4行精确匹配 "number 709133730 RT0001" 模式
                            if 'number ' in fourth_line:
                                # 使用正则表达式精确匹配 "number 709133730 RT0001"
                                # 修改模式：9位数字 + 空格 + RT + 4位数字
                                tax_id_match = re.search(r'number\s+(\d{9}\s+RT\d{4})', fourth_line)
                                if tax_id_match:
                                    vendor_tax_id = tax_id_match.group(1).strip()
                                    # 清理多余空格
                                    vendor_tax_id = ' '.join(vendor_tax_id.split())

                            # 备用方法：从第3行查找 "number" 后的内容
                            if not vendor_tax_id and 'number' in third_line:
                                number_part = third_line.split('number')[1].strip()
                                # 查找以数字开头，可能包含字母的组合
                                number_match = re.search(r'^(\d+[A-Za-z0-9\s]*)', number_part)
                                if number_match:
                                    potential_id = number_match.group(1).strip()
                                    # 确保既包含数字又包含字母
                                    if re.search(r'\d', potential_id) and re.search(r'[A-Za-z]', potential_id):
                                        vendor_tax_id = potential_id

                            # 最后备用：在第3-4行中查找 "709133730 RT0001" 格式
                            if not vendor_tax_id:
                                for check_line in [third_line, fourth_line]:
                                    # 查找特定格式：9位数字 + 空格 + RT + 4位数字
                                    specific_match = re.search(r'709133730\s+RT0001', check_line)
                                    if specific_match:
                                        vendor_tax_id = '709133730 RT0001'
                                        break
                                    # 或者通用格式：9位数字 + 空格 + RT + 4位数字
                                    general_match = re.search(r'(\d{9}\s+RT\d{4})', check_line)
                                    if general_match:
                                        vendor_tax_id = general_match.group(1).strip()
                                        break

                        result['vendor_tax_id'] = vendor_tax_id

                break

        if not need_support_found:
            result['processing_errors'].append("未找到 'Need support' 行")

    except Exception as e:
        result['processing_errors'].append(f"处理过程中出错: {str(e)}")

    return result


def load_field_mapping_config():
    """加载字段映射配置"""
    import json

    config_file = "field_mapping_config.json"
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except Exception as e:
        print(f"❌ 加载配置文件失败: {e}")
        # 返回默认配置
        return {
            "template_file": "Template/导出模板.xlsx",
            "start_row": 5,
            "header_row": 1,
            "sheet_name": "Sheet1",
            "field_mapping": {
                "invoice_number": "M",
                "our_company_address": "AB",
                "our_tax_id": "AC",
                "invoice_date": "L",
                "net_amount": "AQ",
                "tax_rate": "AO",
                "tax_amount": "AP",
                "total_amount": "AR",
                "currency": "AA",
                "vendor_name": "V",
                "vendor_address": "X",
                "vendor_tax_id": "W"
            }
        }


def save_with_template_mapping(df, template_file, output_file):
    """使用模板文件并保持格式，将字段映射到指定的列，从第5行开始插入数据"""
    from openpyxl import load_workbook

    # 加载配置
    config = load_field_mapping_config()
    field_mapping = config.get('field_mapping', {})
    start_row = config.get('start_row', 5)

    try:
        print(f"🔍 尝试加载模板文件: {template_file}")

        # 检查模板文件是否存在
        if not template_file.exists():
            raise FileNotFoundError(f"模板文件不存在: {template_file}")

        # 加载模板文件
        wb = load_workbook(template_file)
        ws = wb.active
        print(f"✅ 模板文件加载成功: {template_file}")
        print(f"📊 工作表: {ws.title}")
        print(f"🔢 从第 {start_row} 行开始写入 {len(df)} 行数据")
        print(f"📝 字段映射: {field_mapping}")

        # 字段映射：源字段 -> 目标列
        # invoice_number → M, our_company_address → AB, our_tax_id → AC, invoice_date → L
        # net_amount → AQ, tax_rate → AO, tax_amount → AP, total_amount → AR
        # currency → AA, vendor_name → V, vendor_address → X, vendor_tax_id → W

        # 从配置的起始行开始写入数据
        for idx, row in df.iterrows():
            current_row = start_row + idx

            # 首先确保保留关键字段，特别是filename
            try:
                # 如果需要，将filename保存到一个临时列（比如AY列），这样后端读取时可以使用
                if 'filename' in row and pd.notna(row['filename']):
                    ws[f"AY{current_row}"] = str(row['filename'])
                    print(f"💾 第 {current_row} 行: 保存filename到AY列: {row['filename']}")
            except Exception as filename_error:
                print(f"⚠️ 保存filename失败: {filename_error}")

            # 映射每个字段到对应的列
            for field_name, target_col in field_mapping.items():
                if field_name in row and pd.notna(row[field_name]):
                    try:
                        col_cell = f"{target_col}{current_row}"
                        value = row[field_name]

                        # 处理特殊字段类型
                        if field_name in ['invoice_date']:
                            # 日期格式化
                            ws[col_cell] = str(value)
                        elif field_name in ['net_amount', 'tax_rate', 'tax_amount', 'total_amount']:
                            # 数值格式
                            try:
                                ws[col_cell] = float(value)
                            except (ValueError, TypeError):
                                ws[col_cell] = 0.0
                        else:
                            # 文本格式
                            ws[col_cell] = str(value)

                    except Exception as cell_error:
                        print(f"⚠️ 写入 {field_name} 到 {target_col}{current_row} 失败: {cell_error}")

            # O列：默认赋值 "tax invoice"
            try:
                ws[f"O{current_row}"] = "tax invoice"
                print(f"💾 第 {current_row} 行: O列赋值 'tax invoice'")
            except Exception as o_error:
                print(f"⚠️ O列赋值失败: {o_error}")

            # S列：根据地址和货币信息转换为ISO代码
            try:
                iso_code = get_country_iso_code_from_address_and_currency(row)
                ws[f"S{current_row}"] = iso_code

                # 显示详细信息
                vendor_address = row.get('vendor_address', '')
                currency = row.get('currency', '')
                address_info = f"地址: {vendor_address}" if pd.notna(vendor_address) and vendor_address else "无地址"
                currency_info = f"货币: {currency}" if pd.notna(currency) and currency else "无货币"

                print(f"💾 第 {current_row} 行: S列赋值ISO代码 '{iso_code}' ({address_info}, {currency_info})")
            except Exception as s_error:
                print(f"⚠️ S列赋值失败: {s_error}")
                try:
                    ws[f"S{current_row}"] = "US"  # 出错时使用默认值
                except:
                    pass

            print(f"✅ 第 {current_row} 行数据已写入（含O列和S列）")

        # 保存文件
        print(f"💾 正在保存文件: {output_file}")
        wb.save(output_file)
        print(f"✅ 模板保存成功: {output_file}")

        return True

    except Exception as e:
        print(f"❌ 模板保存失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def get_country_iso_code_from_address_and_currency(row):
    """根据地址和货币信息确定ISO国家代码"""
    # 首先尝试从vendor_address中提取国家信息
    vendor_address = row.get('vendor_address', '')
    if pd.notna(vendor_address) and vendor_address:
        iso_code = get_country_iso_code(vendor_address)
        if iso_code != "US":  # 如果识别出非美国国家，直接返回
            return iso_code

    # 如果地址中没有识别出具体国家，尝试从currency字段判断
    currency = row.get('currency', '')
    if pd.notna(currency):
        currency_str = str(currency).strip().upper()

        # 货币到ISO代码的映射
        currency_mapping = {
            "USD": "US",
            "$": "US",  # 美元符号
            "EUR": "DE",  # 欧元默认德国（欧元区）
            "€": "DE",   # 欧元符号
            "GBP": "GB",
            "£": "GB",   # 英镑符号
            "CNY": "CN",
            "RMB": "CN",
            "¥": "CN",   # 人民币符号
            "JPY": "JP",
            "¥": "JP",   # 日元符号
            "AUD": "AU",
            "CAD": "CA",
            "CHF": "CH",
            "SEK": "SE",  # 瑞典克朗
            "DKK": "DK",  # 丹麦克朗
            "NOK": "NO",  # 挪威克朗
            "INR": "IN",  # 印度卢比
            "KRW": "KR",  # 韩元
            "SGD": "SG",  # 新加坡元
            "HKD": "HK",  # 港币
            "MYR": "MY",  # 马来西亚林吉特
            "THB": "TH",  # 泰铢
            "PHP": "PHP", # 菲律宾比索
            "VND": "VN",  # 越南盾
            "TRY": "TR",  # 土耳其里拉
            "ILS": "IL",  # 以色列新谢克尔
            "AED": "AE",  # 阿联酋迪拉姆
            "SAR": "SA",  # 沙特里亚尔
            "NZD": "NZ",  # 新西兰元
            "RUB": "RU",  # 俄罗斯卢布
            "BRL": "BR",  # 巴西雷亚尔
            "ARS": "AR",  # 阿根廷比索
            "CLP": "CL",  # 智利比索
            "ZAR": "ZA",  # 南非兰特
            "EGP": "EG",  # 埃及镑
            "NGN": "NG",  # 尼日利亚奈拉
            "KES": "KE",  # 肯尼亚先令
        }

        if currency_str in currency_mapping:
            iso_code = currency_mapping[currency_str]
            print(f"💰 基于货币 {currency_str} 识别国家代码: {iso_code}")
            return iso_code

    # 默认返回US
    return "US"


def get_country_iso_code(country_name):
    """将国家名称转换为ISO 2位国家代码"""
    if pd.isna(country_name):
        return "US"  # 默认值

    country_mapping = {
        # 欧洲
        "United Kingdom": "GB",
        "UK": "GB",
        "England": "GB",
        "Scotland": "GB",
        "Wales": "GB",
        "Northern Ireland": "GB",
        "Great Britain": "GB",
        "Britain": "GB",
        "Sweden": "SE",
        "Swedish": "SE",
        "Sverige": "SE",  # 瑞典语
        "Stockholm": "SE",  # 首都
        "Gothenburg": "SE",
        "Malmo": "SE",
        "Germany": "DE",
        "France": "FR",
        "Italy": "IT",
        "Spain": "ES",
        "Netherlands": "NL",
        "Belgium": "BE",
        "Poland": "PL",
        "Denmark": "DK",
        "Norway": "NO",
        "Finland": "FI",
        "Austria": "AT",
        "Switzerland": "CH",
        "Ireland": "IE",
        "Portugal": "PT",
        "Czech Republic": "CZ",
        "Hungary": "HU",
        "Romania": "RO",
        "Bulgaria": "BG",
        "Greece": "GR",
        "Croatia": "HR",
        "Slovakia": "SK",
        "Slovenia": "SI",
        "Estonia": "EE",
        "Latvia": "LV",
        "Lithuania": "LT",
        "Luxembourg": "LU",
        "Malta": "MT",
        "Cyprus": "CY",

        # 北美洲
        "United States": "US",
        "USA": "US",
        "America": "US",
        "Canada": "CA",
        "Mexico": "MX",

        # 亚洲
        "China": "CN",
        "PRC": "CN",
        "People's Republic of China": "CN",
        "Japan": "JP",
        "South Korea": "KR",
        "Korea": "KR",
        "Singapore": "SG",
        "Hong Kong": "HK",
        "Taiwan": "TW",
        "India": "IN",
        "Indonesia": "ID",
        "Thailand": "TH",
        "Malaysia": "MY",
        "Philippines": "PH",
        "Vietnam": "VN",
        "Turkey": "TR",
        "Israel": "IL",
        "UAE": "AE",
        "United Arab Emirates": "AE",
        "Saudi Arabia": "SA",

        # 大洋洲
        "Australia": "AU",
        "New Zealand": "NZ",

        # 其他
        "Russia": "RU",
        "Brazil": "BR",
        "Argentina": "AR",
        "Chile": "CL",
        "South Africa": "ZA",
        "Egypt": "EG",
        "Nigeria": "NG",
        "Kenya": "KE"
    }

    # 尝试直接匹配
    country_str = str(country_name).strip()
    if country_str in country_mapping:
        return country_mapping[country_str]

    # 尝试模糊匹配（不区分大小写，去除标点）
    import re
    clean_country = re.sub(r'[^\w\s]', '', country_str.lower())

    for key, code in country_mapping.items():
        clean_key = re.sub(r'[^\w\s]', '', key.lower())
        if clean_key in clean_country or clean_country in clean_key:
            return code

    # 如果都匹配不到，根据关键词猜测
    lower_country = country_str.lower()
    if any(word in lower_country for word in ['uk', 'britain', 'england', 'scotland', 'wales']):
        return "GB"
    elif any(word in lower_country for word in ['china', 'chinese']):
        return "CN"
    elif any(word in lower_country for word in ['america', 'usa', 'states']):
        return "US"
    elif any(word in lower_country for word in ['australia', 'australian']):
        return "AU"
    elif any(word in lower_country for word in ['canada', 'canadian']):
        return "CA"

    # 默认返回US
    print(f"⚠️ 无法识别国家: {country_name}，使用默认值US")
    return "US"


def main(progress_callback=None, file_processed_callback=None):
    """主函数：处理所有/debug_txt下的文件"""
    print("🏢 [FORMAL] 全OU公司Klarna发票数据提取器")
    print("=" * 60)
    print("⚠️  正式版本：支持所有8种OU公司类型，包含所有修复")
    print()

    debug_txt_path = Path("./debug_txt")
    if not debug_txt_path.exists():
        print(f"❌ 错误: 找不到文件夹 {debug_txt_path}")
        return

    results = []

    # 获取所有txt文件
    txt_files = list(debug_txt_path.glob("*.txt"))
    print(f"📄 找到 {len(txt_files)} 个txt文件")

    if not txt_files:
        print("❌ 未找到任何txt文件")
        return

    total_files = len(txt_files)

    # 处理每个文件
    for i, file_path in enumerate(txt_files, 1):
        # Update progress
        if progress_callback:
            try:
                progress_callback(i, total_files)
            except Exception:
                pass

        print(f"处理: {file_path.name}")

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            # 检测OU公司并选择对应的提取函数
            if len(lines) >= 8:
                line_8 = lines[7].strip()
                separator = "     "
                if separator in line_8:
                    ou_company = line_8.split(separator)[0].strip()
                else:
                    ou_company = "未知公司"
            else:
                ou_company = "未知公司"

            company_type = detect_company_type(ou_company)

            # 根据公司类型选择提取函数
            if company_type == "AUSTRALIA":
                result = extract_shein_australia_data(lines)
            elif company_type == "UK":
                result = extract_shein_uk_data(lines)
            elif company_type == "IRELAND":
                result = extract_infinite_styles_ireland_data(lines)
            elif company_type == "TOWERS":
                result = extract_infinite_towers_data(lines)
            elif company_type == "STYLES_SERVICES":
                result = extract_infinite_styles_services_data(lines)
            elif company_type == "CORPORATION":
                result = extract_shein_corporation_data(lines)
            elif company_type == "US_SERVICES":
                result = extract_shein_us_services_data(lines)
            elif company_type == "CANADA":
                result = extract_shein_canada_data(lines)
            else:
                # 不支持的公司类型，创建基础记录
                result = {
                    'invoice_number': '',
                    'our_company_name': ou_company,
                    'our_company_address': '',
                    'our_tax_id': '',
                    'invoice_date': '',
                    'net_amount': '',
                    'tax_rate': '',
                    'tax_amount': '',
                    'total_amount': '',
                    'currency': '',
                    'vendor_name': '',
                    'vendor_address': '',
                    'vendor_tax_id': '',
                    'filename': file_path.name,
                    'processing_errors': [f"暂不支持 {ou_company} 的提取逻辑"]
                }

            result['filename'] = file_path.name
            results.append(result)

            # 实时回调：通知前端有新文件处理完成
            if file_processed_callback:
                try:
                    file_processed_callback(result)
                except Exception as callback_error:
                    print(f"⚠️ 文件处理回调失败: {callback_error}")

        except Exception as e:
            print(f"   ❌ 处理 {file_path.name} 时出错: {str(e)}")
            error_result = {
                'invoice_number': '',
                'our_company_name': '处理错误',
                'our_company_address': '',
                'our_tax_id': '',
                'invoice_date': '',
                'net_amount': '',
                'tax_rate': '',
                'tax_amount': '',
                'total_amount': '',
                'currency': '',
                'vendor_name': '',
                'vendor_address': '',
                'vendor_tax_id': '',
                'filename': file_path.name,
                'processing_errors': [f"文件读取错误: {str(e)}"]
            }
            results.append(error_result)

            # 实时回调：通知前端有新文件处理完成（即使是错误）
            if file_processed_callback:
                try:
                    file_processed_callback(error_result)
                except Exception as callback_error:
                    print(f"⚠️ 错误文件处理回调失败: {callback_error}")

    if not results:
        print("❌ 没有成功处理任何文件")
        return

    # 创建DataFrame
    df = pd.DataFrame(results)

    # 数据清理
    def clean_for_excel(value):
        try:
            if value is None or (hasattr(value, '__len__') and len(value) == 0):
                return ""
            if not isinstance(value, str):
                value = str(value)
        except:
            return ""

        cleaned = value
        # 清理33种控制字符，保留制表符、换行符和回车符
        for i in range(32):
            if i not in (9, 10, 13):
                cleaned = cleaned.replace(chr(i), '')
        # 标准化空格
        cleaned = ' '.join(cleaned.split())
        # 截断过长的内容
        if len(cleaned) > 32700:
            cleaned = cleaned[:32700] + "..."
        return cleaned

    df_clean = df.copy()
    for col in df_clean.columns:
        if df_clean[col].dtype == 'object':
            df_clean[col] = df_clean[col].apply(clean_for_excel)

    # 保存到Excel - 使用模板并映射字段
    output_file = "FORMAL_ALL_OU_COMPANIES.xlsx"

    # 从配置文件加载模板路径
    config = load_field_mapping_config()
    template_file = Path(config.get('template_file', 'Template/导出模板.xlsx'))

    export_success = False

    try:
        print(f"📄 模板文件路径: {template_file}")

        # 确保filename字段存在
        if 'filename' not in df_clean.columns:
            print("⚠️ filename列不存在，创建默认值")
            df_clean['filename'] = [f'processed_file_{i+1}.pdf' for i in range(len(df_clean))]

        print(f"📋 filename列示例: {df_clean['filename'].head(5).tolist()}")

        # 检查模板文件是否存在
        if template_file.exists():
            print(f"✅ 找到模板文件: {template_file}")
            export_success = save_with_template_mapping(df_clean, template_file, output_file)
            if not export_success:
                print("❌ 模板导出失败，使用默认方式")
                df_clean.to_excel(output_file, index=False)
                export_success = True
        else:
            print(f"⚠️ 模板文件不存在: {template_file}")
            print("🔄 使用默认方式保存...")
            df_clean.to_excel(output_file, index=False)
            export_success = True

        print(f"\n✅ 成功生成文件: {output_file}")
        print(f"📊 处理了 {len(df)} 个文件")

    except Exception as e:
        print(f"❌ 导出过程发生错误: {e}")
        import traceback
        traceback.print_exc()
        try:
            df_clean.to_excel(output_file, index=False)
            export_success = True
            print("✅ 降级保存成功")
        except Exception as final_error:
            print(f"❌ 最终保存失败: {final_error}")
            return False

    return export_success


if __name__ == "__main__":
    main()